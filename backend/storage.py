import os
import io
import logging
import requests as http_requests
from ddgs import DDGS

# ===== Object Storage =====
STORAGE_URL = "https://integrations.emergentagent.com/objstore/api/v1/storage"
EMERGENT_KEY = os.environ.get("EMERGENT_LLM_KEY")
APP_NAME = "openwebui-clone"
_storage_key = None


def init_storage():
    global _storage_key
    if _storage_key:
        return _storage_key
    resp = http_requests.post(f"{STORAGE_URL}/init", json={"emergent_key": EMERGENT_KEY}, timeout=30)
    resp.raise_for_status()
    _storage_key = resp.json()["storage_key"]
    return _storage_key


def put_object(path: str, data: bytes, content_type: str) -> dict:
    key = init_storage()
    resp = http_requests.put(
        f"{STORAGE_URL}/objects/{path}",
        headers={"X-Storage-Key": key, "Content-Type": content_type},
        data=data, timeout=120
    )
    resp.raise_for_status()
    return resp.json()


def get_object(path: str):
    key = init_storage()
    resp = http_requests.get(
        f"{STORAGE_URL}/objects/{path}",
        headers={"X-Storage-Key": key}, timeout=60
    )
    resp.raise_for_status()
    return resp.content, resp.headers.get("Content-Type", "application/octet-stream")


# ===== File Processing =====
IMAGE_EXTENSIONS = {"jpg", "jpeg", "png", "gif", "webp", "svg", "bmp", "ico"}
TEXT_EXTENSIONS = {"txt", "md", "csv", "json", "xml", "yaml", "yml", "html", "css", "js", "jsx", "ts", "tsx", "py", "java", "c", "cpp", "h", "go", "rs", "rb", "php", "sh", "bash", "sql", "r", "swift", "kt", "toml", "ini", "cfg", "log", "env"}
MAX_FILE_SIZE = 20 * 1024 * 1024  # 20MB


def extract_text_from_file(data: bytes, filename: str, content_type: str) -> str:
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    try:
        if ext == "pdf" or content_type == "application/pdf":
            from PyPDF2 import PdfReader
            reader = PdfReader(io.BytesIO(data))
            text = "\n".join(page.extract_text() or "" for page in reader.pages)
            return text.strip()[:50000]
        elif ext == "docx" or content_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
            from docx import Document
            doc = Document(io.BytesIO(data))
            text = "\n".join(p.text for p in doc.paragraphs)
            return text.strip()[:50000]
        elif ext == "xlsx" or content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(data), read_only=True)
            rows = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                rows.append(f"--- Sheet: {sheet} ---")
                for row in ws.iter_rows(values_only=True):
                    rows.append(",".join(str(c) if c is not None else "" for c in row))
            return "\n".join(rows)[:50000]
        elif ext in TEXT_EXTENSIONS:
            return data.decode("utf-8", errors="replace")[:50000]
        else:
            return ""
    except Exception as e:
        return f"[Error extracting text: {str(e)}]"


def is_image_file(filename: str, content_type: str) -> bool:
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    return ext in IMAGE_EXTENSIONS or (content_type or "").startswith("image/")


# ===== Web Search =====
def web_search(query: str, max_results: int = 5) -> str:
    try:
        with DDGS() as ddgs:
            results = list(ddgs.text(query, max_results=max_results))
        if not results:
            return ""
        formatted = []
        for i, r in enumerate(results, 1):
            title = r.get("title", "")
            body = r.get("body", "")
            href = r.get("href", "")
            formatted.append(f"[{i}] {title}\n{body}\nSource: {href}")
        return "\n\n".join(formatted)
    except Exception as e:
        logging.getLogger(__name__).warning(f"Web search failed: {e}")
        return ""
