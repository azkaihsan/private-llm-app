from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request, UploadFile, File, Query
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import Response
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional
import uuid
import base64
import io
from datetime import datetime, timezone, timedelta
from emergentintegrations.llm.chat import LlmChat, UserMessage, ImageContent
import litellm
import jwt
from passlib.context import CryptContext
import requests as http_requests
from ddgs import DDGS


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# ===== Auth Config =====
JWT_SECRET = os.environ.get("JWT_SECRET", "openwebui-clone-secret-key-change-in-production")
JWT_ALGORITHM = "HS256"
JWT_EXPIRY_HOURS = 72
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer(auto_error=False)

def hash_password(password: str) -> str:
    return pwd_context.hash(password)

def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)

def create_token(user_id: str, role: str) -> str:
    payload = {
        "sub": user_id,
        "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRY_HOURS),
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def decode_token(token: str) -> dict:
    try:
        return jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    if not credentials:
        raise HTTPException(status_code=401, detail="Not authenticated")
    payload = decode_token(credentials.credentials)
    user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    return user

async def get_admin_user(user=Depends(get_current_user)):
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return user

# Optional auth - returns user or None
async def get_optional_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    if not credentials:
        return None
    try:
        payload = decode_token(credentials.credentials)
        user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
        return user
    except Exception:
        return None


# ===== Object Storage =====
STORAGE_URL = "https://integrations.emergentagent.com/objstore/api/v1/storage"
EMERGENT_KEY = os.environ.get("EMERGENT_LLM_KEY")
APP_NAME = "openwebui-clone"
_storage_key = None

def init_storage():
    global _storage_key
    if _storage_key:
        return _storage_key
    resp = http_requests.post(f"{STORAGE_URL}/init", json={"emergent_key": EMERGENT_KEY}, timeout=30)
    resp.raise_for_status()
    _storage_key = resp.json()["storage_key"]
    return _storage_key

def put_object(path: str, data: bytes, content_type: str) -> dict:
    key = init_storage()
    resp = http_requests.put(
        f"{STORAGE_URL}/objects/{path}",
        headers={"X-Storage-Key": key, "Content-Type": content_type},
        data=data, timeout=120
    )
    resp.raise_for_status()
    return resp.json()

def get_object(path: str):
    key = init_storage()
    resp = http_requests.get(
        f"{STORAGE_URL}/objects/{path}",
        headers={"X-Storage-Key": key}, timeout=60
    )
    resp.raise_for_status()
    return resp.content, resp.headers.get("Content-Type", "application/octet-stream")


# ===== File Processing Utilities =====
IMAGE_EXTENSIONS = {"jpg", "jpeg", "png", "gif", "webp", "svg", "bmp", "ico"}
TEXT_EXTENSIONS = {"txt", "md", "csv", "json", "xml", "yaml", "yml", "html", "css", "js", "jsx", "ts", "tsx", "py", "java", "c", "cpp", "h", "go", "rs", "rb", "php", "sh", "bash", "sql", "r", "swift", "kt", "toml", "ini", "cfg", "log", "env"}
MAX_FILE_SIZE = 20 * 1024 * 1024  # 20MB

def extract_text_from_file(data: bytes, filename: str, content_type: str) -> str:
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    try:
        if ext == "pdf" or content_type == "application/pdf":
            from PyPDF2 import PdfReader
            reader = PdfReader(io.BytesIO(data))
            text = "\n".join(page.extract_text() or "" for page in reader.pages)
            return text.strip()[:50000]
        elif ext == "docx" or content_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
            from docx import Document
            doc = Document(io.BytesIO(data))
            text = "\n".join(p.text for p in doc.paragraphs)
            return text.strip()[:50000]
        elif ext == "xlsx" or content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(data), read_only=True)
            rows = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                rows.append(f"--- Sheet: {sheet} ---")
                for row in ws.iter_rows(values_only=True):
                    rows.append(",".join(str(c) if c is not None else "" for c in row))
            return "\n".join(rows)[:50000]
        elif ext in TEXT_EXTENSIONS:
            return data.decode("utf-8", errors="replace")[:50000]
        else:
            return ""
    except Exception as e:
        return f"[Error extracting text: {str(e)}]"

def is_image_file(filename: str, content_type: str) -> bool:
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    return ext in IMAGE_EXTENSIONS or (content_type or "").startswith("image/")


# ===== Web Search =====
def web_search(query: str, max_results: int = 5) -> str:
    """Search the web using DuckDuckGo and return formatted results."""
    try:
        with DDGS() as ddgs:
            results = list(ddgs.text(query, max_results=max_results))
        if not results:
            return ""
        formatted = []
        for i, r in enumerate(results, 1):
            title = r.get("title", "")
            body = r.get("body", "")
            href = r.get("href", "")
            formatted.append(f"[{i}] {title}\n{body}\nSource: {href}")
        return "\n\n".join(formatted)
    except Exception as e:
        logging.getLogger(__name__).warning(f"Web search failed: {e}")
        return ""


# Define Models
class StatusCheck(BaseModel):
    model_config = ConfigDict(extra="ignore")  # Ignore MongoDB's _id field
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    client_name: str
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class StatusCheckCreate(BaseModel):
    client_name: str

# Add your routes to the router instead of directly to app
@api_router.get("/")
async def root():
    return {"message": "Hello World"}

@api_router.post("/status", response_model=StatusCheck)
async def create_status_check(input: StatusCheckCreate):
    status_dict = input.model_dump()
    status_obj = StatusCheck(**status_dict)
    
    # Convert to dict and serialize datetime to ISO string for MongoDB
    doc = status_obj.model_dump()
    doc['timestamp'] = doc['timestamp'].isoformat()
    
    _ = await db.status_checks.insert_one(doc)
    return status_obj

@api_router.get("/status", response_model=List[StatusCheck])
async def get_status_checks():
    # Exclude MongoDB's _id field from the query results
    status_checks = await db.status_checks.find({}, {"_id": 0}).to_list(1000)
    
    # Convert ISO string timestamps back to datetime objects
    for check in status_checks:
        if isinstance(check['timestamp'], str):
            check['timestamp'] = datetime.fromisoformat(check['timestamp'])
    
    return status_checks

# ===== Auth Endpoints =====

class SignupRequest(BaseModel):
    name: str
    email: str
    password: str

class LoginRequest(BaseModel):
    email: str
    password: str

@api_router.post("/auth/signup")
async def signup(data: SignupRequest):
    existing = await db.users.find_one({"email": data.email.lower()})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")

    # First user becomes admin
    user_count = await db.users.count_documents({})
    role = "admin" if user_count == 0 else "user"

    user = {
        "id": str(uuid.uuid4()),
        "name": data.name,
        "email": data.email.lower(),
        "password_hash": hash_password(data.password),
        "role": role,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.users.insert_one({**user, "_id": user["id"]})
    token = create_token(user["id"], user["role"])
    return {
        "token": token,
        "user": {"id": user["id"], "name": user["name"], "email": user["email"], "role": user["role"]},
    }

@api_router.post("/auth/login")
async def login(data: LoginRequest):
    user = await db.users.find_one({"email": data.email.lower()})
    if not user or not verify_password(data.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")

    token = create_token(user["id"], user["role"])
    return {
        "token": token,
        "user": {"id": user["id"], "name": user["name"], "email": user["email"], "role": user["role"]},
    }

@api_router.get("/auth/me")
async def get_me(user=Depends(get_current_user)):
    return user

# ===== Admin: User Management =====

class AdminCreateUser(BaseModel):
    name: str
    email: str
    password: str
    role: str = "user"

class AdminUpdateUser(BaseModel):
    name: Optional[str] = None
    email: Optional[str] = None
    role: Optional[str] = None
    password: Optional[str] = None

@api_router.get("/admin/users")
async def admin_list_users(admin=Depends(get_admin_user)):
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).sort("created_at", -1).to_list(500)
    return users

@api_router.post("/admin/users")
async def admin_create_user(data: AdminCreateUser, admin=Depends(get_admin_user)):
    existing = await db.users.find_one({"email": data.email.lower()})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    if data.role not in ("admin", "user"):
        raise HTTPException(status_code=400, detail="Role must be 'admin' or 'user'")

    user = {
        "id": str(uuid.uuid4()),
        "name": data.name,
        "email": data.email.lower(),
        "password_hash": hash_password(data.password),
        "role": data.role,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.users.insert_one({**user, "_id": user["id"]})
    return {"id": user["id"], "name": user["name"], "email": user["email"], "role": user["role"], "created_at": user["created_at"]}

@api_router.put("/admin/users/{user_id}")
async def admin_update_user(user_id: str, data: AdminUpdateUser, admin=Depends(get_admin_user)):
    update = {}
    if data.name is not None:
        update["name"] = data.name
    if data.email is not None:
        existing = await db.users.find_one({"email": data.email.lower(), "id": {"$ne": user_id}})
        if existing:
            raise HTTPException(status_code=400, detail="Email already in use")
        update["email"] = data.email.lower()
    if data.role is not None:
        if data.role not in ("admin", "user"):
            raise HTTPException(status_code=400, detail="Role must be 'admin' or 'user'")
        update["role"] = data.role
    if data.password is not None:
        update["password_hash"] = hash_password(data.password)

    if not update:
        raise HTTPException(status_code=400, detail="No fields to update")

    result = await db.users.update_one({"id": user_id}, {"$set": update})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {"status": "ok"}

@api_router.delete("/admin/users/{user_id}")
async def admin_delete_user(user_id: str, admin=Depends(get_admin_user)):
    if admin["id"] == user_id:
        raise HTTPException(status_code=400, detail="Cannot delete yourself")
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    # Also delete user's chats and messages
    user_chats = await db.chats.find({"user_id": user_id}, {"id": 1}).to_list(1000)
    for chat in user_chats:
        await db.messages.delete_many({"chat_id": chat["id"]})
    await db.chats.delete_many({"user_id": user_id})
    return {"status": "ok"}

# ===== File Upload/Download Endpoints =====

@api_router.post("/files/upload")
async def upload_file(file: UploadFile = File(...), user=Depends(get_current_user)):
    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename provided")
    data = await file.read()
    if len(data) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="File too large (max 20MB)")

    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else "bin"
    file_id = str(uuid.uuid4())
    storage_path = f"{APP_NAME}/uploads/{user['id']}/{file_id}.{ext}"
    content_type = file.content_type or "application/octet-stream"

    result = put_object(storage_path, data, content_type)

    file_doc = {
        "id": file_id,
        "storage_path": result["path"],
        "original_filename": file.filename,
        "content_type": content_type,
        "size": len(data),
        "user_id": user["id"],
        "is_image": is_image_file(file.filename, content_type),
        "is_deleted": False,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.files.insert_one({**file_doc, "_id": file_id})
    return {k: v for k, v in file_doc.items() if k != "_id"}

@api_router.get("/files/{file_id}")
async def download_file(file_id: str, auth: str = Query(None), user=Depends(get_optional_user)):
    # Support query param auth for img tags
    if not user and auth:
        try:
            payload = decode_token(auth)
            user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
        except Exception:
            pass
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")

    record = await db.files.find_one({"id": file_id, "is_deleted": False}, {"_id": 0})
    if not record:
        raise HTTPException(status_code=404, detail="File not found")

    data, ct = get_object(record["storage_path"])
    return Response(content=data, media_type=record.get("content_type", ct))

# Include the router in the main app - moved to after all routes are defined

# ===== Chat Models =====
ALL_MODELS = {
    "openai": [
        {"id": "gpt-4o", "name": "GPT-4o", "provider": "openai"},
        {"id": "gpt-4.1-mini", "name": "GPT-4.1 mini", "provider": "openai"},
        {"id": "gpt-5.1", "name": "GPT-5.1", "provider": "openai"},
        {"id": "gpt-5-mini", "name": "GPT-5 mini", "provider": "openai"},
    ],
    "anthropic": [
        {"id": "claude-sonnet-4-5-20250929", "name": "Claude Sonnet 4.5", "provider": "anthropic"},
        {"id": "claude-4-sonnet-20250514", "name": "Claude 4 Sonnet", "provider": "anthropic"},
    ],
    "gemini": [
        {"id": "gemini-2.5-flash", "name": "Gemini 2.5 Flash", "provider": "gemini"},
        {"id": "gemini-2.5-pro", "name": "Gemini 2.5 Pro", "provider": "gemini"},
    ],
    "deepseek": [
        {"id": "deepseek/deepseek-chat", "name": "DeepSeek V3", "provider": "deepseek"},
        {"id": "deepseek/deepseek-reasoner", "name": "DeepSeek R1", "provider": "deepseek"},
    ],
    "qwen": [
        {"id": "openai/qwen-max", "name": "Qwen Max", "provider": "qwen"},
        {"id": "openai/qwen-plus", "name": "Qwen Plus", "provider": "qwen"},
        {"id": "openai/qwen-turbo", "name": "Qwen Turbo", "provider": "qwen"},
    ],
    "grok": [
        {"id": "xai/grok-3", "name": "Grok 3", "provider": "grok"},
        {"id": "xai/grok-3-mini", "name": "Grok 3 Mini", "provider": "grok"},
    ],
    "perplexity": [
        {"id": "perplexity/sonar-pro", "name": "Sonar Pro", "provider": "perplexity"},
        {"id": "perplexity/sonar", "name": "Sonar", "provider": "perplexity"},
        {"id": "perplexity/sonar-reasoning-pro", "name": "Sonar Reasoning Pro", "provider": "perplexity"},
    ],
    "bedrock": [
        {"id": "bedrock/anthropic.claude-3-5-sonnet-20241022-v2:0", "name": "Claude 3.5 Sonnet (Bedrock)", "provider": "bedrock"},
        {"id": "bedrock/amazon.nova-pro-v1:0", "name": "Amazon Nova Pro", "provider": "bedrock"},
        {"id": "bedrock/amazon.nova-lite-v1:0", "name": "Amazon Nova Lite", "provider": "bedrock"},
    ],
    "openai_compatible": [],
}

# Providers that use Emergent key (only openai, anthropic, gemini)
EMERGENT_PROVIDERS = {"openai", "anthropic", "gemini"}

# Provider base URLs for non-Emergent providers
PROVIDER_BASE_URLS = {
    "deepseek": "https://api.deepseek.com/v1",
    "qwen": "https://dashscope-intl.aliyuncs.com/compatible-mode/v1",
    "grok": "https://api.x.ai/v1",
    "perplexity": "https://api.perplexity.ai",
}

FLAT_MODELS = [m for models in ALL_MODELS.values() for m in models]
MODEL_PROVIDER_MAP = {m["id"]: m["provider"] for m in FLAT_MODELS}

class ChatCreate(BaseModel):
    title: str = ""
    model: str = "gpt-4o"

class ChatUpdate(BaseModel):
    title: str

class MessageCreate(BaseModel):
    content: str
    file_ids: Optional[List[str]] = None
    temperature: Optional[float] = None
    max_tokens: Optional[int] = None
    top_p: Optional[float] = None

# ===== Connection & Model Endpoints =====

@api_router.get("/connections")
async def get_connections(admin=Depends(get_admin_user)):
    doc = await db.connections.find_one({"_id": "global"}, {"_id": 0})
    if not doc:
        default_key = os.environ.get("EMERGENT_LLM_KEY", "")
        return {
            "providers": {
                "openai": {"enabled": True, "apiKey": default_key, "name": "OpenAI", "useEmergentKey": True},
                "anthropic": {"enabled": True, "apiKey": default_key, "name": "Anthropic", "useEmergentKey": True},
                "gemini": {"enabled": True, "apiKey": default_key, "name": "Google Gemini", "useEmergentKey": True},
                "deepseek": {"enabled": False, "apiKey": "", "name": "DeepSeek", "useEmergentKey": False, "baseUrl": "https://api.deepseek.com/v1"},
                "qwen": {"enabled": False, "apiKey": "", "name": "Qwen", "useEmergentKey": False, "baseUrl": "https://dashscope-intl.aliyuncs.com/compatible-mode/v1"},
                "grok": {"enabled": False, "apiKey": "", "name": "Grok (xAI)", "useEmergentKey": False, "baseUrl": "https://api.x.ai/v1"},
                "perplexity": {"enabled": False, "apiKey": "", "name": "Perplexity", "useEmergentKey": False, "baseUrl": "https://api.perplexity.ai"},
                "bedrock": {"enabled": False, "apiKey": "", "name": "Amazon Bedrock", "useEmergentKey": False, "awsRegion": "us-east-1", "awsAccessKey": "", "awsSecretKey": ""},
                "openai_compatible": {"enabled": False, "apiKey": "", "name": "OpenAI Compatible", "useEmergentKey": False, "baseUrl": "", "customModels": ""},
            },
            "defaultModel": "gpt-4o",
            "modelParams": {"temperature": 0.7, "maxTokens": 4096, "topP": 1.0},
            "disabledModels": [],
        }
    return doc

@api_router.put("/connections")
async def update_connections(data: dict, admin=Depends(get_admin_user)):
    await db.connections.update_one(
        {"_id": "global"},
        {"$set": data},
        upsert=True
    )
    return {"status": "ok"}

class TestConnectionRequest(BaseModel):
    provider: str
    apiKey: str

@api_router.post("/connections/test")
async def test_connection(data: TestConnectionRequest, admin=Depends(get_admin_user)):
    try:
        llm = LlmChat(
            api_key=data.apiKey,
            session_id=f"test-{uuid.uuid4()}",
            system_message="Reply with exactly: CONNECTION_OK"
        ).with_model(data.provider, None)
        result = await llm.send_message(UserMessage(text="Say CONNECTION_OK"))  # noqa: F841
        return {"status": "ok", "message": "Connection successful"}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@api_router.get("/models")
async def get_models():
    conn = await db.connections.find_one({"_id": "global"}, {"_id": 0})
    if not conn:
        # Use default configuration - only enabled providers (openai, anthropic, gemini)
        providers = {
            "openai": {"enabled": True},
            "anthropic": {"enabled": True},
            "gemini": {"enabled": True},
            "deepseek": {"enabled": False},
            "qwen": {"enabled": False},
            "grok": {"enabled": False},
            "perplexity": {"enabled": False},
            "bedrock": {"enabled": False},
            "openai_compatible": {"enabled": False},
        }
        disabled = set()
    else:
        providers = conn.get("providers", {})
        disabled = set(conn.get("disabledModels", []))
    
    result = []
    for provider_key, models in ALL_MODELS.items():
        if provider_key == "openai_compatible":
            continue
        prov_config = providers.get(provider_key, {})
        if prov_config.get("enabled", False if provider_key not in EMERGENT_PROVIDERS else True):
            for m in models:
                model_entry = {**m, "enabled": m["id"] not in disabled}
                result.append(model_entry)

    # Add custom OpenAI Compatible models
    compat_config = providers.get("openai_compatible", {})
    if compat_config.get("enabled", False) and compat_config.get("customModels", ""):
        custom_models = [m.strip() for m in compat_config["customModels"].split(",") if m.strip()]
        for cm in custom_models:
            model_entry = {"id": f"openai/{cm}", "name": cm, "provider": "openai_compatible", "enabled": f"openai/{cm}" not in disabled}
            result.append(model_entry)

    return result

# ===== Chat Endpoints =====

@api_router.get("/chats")
async def get_chats(user=Depends(get_current_user)):
    chats = await db.chats.find({"user_id": user["id"], "archived": {"$ne": True}}, {"_id": 0}).sort("created_at", -1).to_list(100)
    return chats

@api_router.post("/chats")
async def create_chat(data: ChatCreate, user=Depends(get_current_user)):
    chat = {
        "id": str(uuid.uuid4()),
        "title": data.title or "New Chat",
        "model": data.model,
        "user_id": user["id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.chats.insert_one({**chat, "_id": chat["id"]})
    return chat

@api_router.get("/chats/archived")
async def get_archived_chats(user=Depends(get_current_user)):
    chats = await db.chats.find({"user_id": user["id"], "archived": True}, {"_id": 0}).sort("created_at", -1).to_list(100)
    return chats

@api_router.get("/chats/{chat_id}")
async def get_chat(chat_id: str, user=Depends(get_current_user)):
    chat = await db.chats.find_one({"id": chat_id, "user_id": user["id"]}, {"_id": 0})
    if not chat:
        raise HTTPException(status_code=404, detail="Chat not found")
    messages = await db.messages.find(
        {"chat_id": chat_id}, {"_id": 0}
    ).sort("timestamp", 1).to_list(1000)
    chat["messages"] = messages
    return chat

@api_router.put("/chats/{chat_id}")
async def update_chat(chat_id: str, data: ChatUpdate, user=Depends(get_current_user)):
    result = await db.chats.update_one({"id": chat_id, "user_id": user["id"]}, {"$set": {"title": data.title}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Chat not found")
    return {"status": "ok"}

@api_router.delete("/chats/{chat_id}")
async def delete_chat(chat_id: str, user=Depends(get_current_user)):
    await db.chats.delete_one({"id": chat_id, "user_id": user["id"]})
    await db.messages.delete_many({"chat_id": chat_id})
    return {"status": "ok"}

@api_router.post("/chats/{chat_id}/messages")
async def send_message(chat_id: str, data: MessageCreate, user=Depends(get_current_user)):
    chat = await db.chats.find_one({"id": chat_id, "user_id": user["id"]}, {"_id": 0})
    if not chat:
        raise HTTPException(status_code=404, detail="Chat not found")

    now_ts = datetime.now(timezone.utc).isoformat()

    # Process file attachments
    attachments = []
    file_text_parts = []
    image_b64_list = []

    if data.file_ids:
        for fid in data.file_ids:
            file_rec = await db.files.find_one({"id": fid, "user_id": user["id"], "is_deleted": False}, {"_id": 0})
            if not file_rec:
                continue
            attachments.append({
                "id": file_rec["id"],
                "filename": file_rec["original_filename"],
                "content_type": file_rec["content_type"],
                "size": file_rec["size"],
                "is_image": file_rec.get("is_image", False),
            })
            try:
                file_data, _ = get_object(file_rec["storage_path"])
                if file_rec.get("is_image", False):
                    b64 = base64.b64encode(file_data).decode("utf-8")
                    mime = file_rec["content_type"] or "image/png"
                    image_b64_list.append({"b64": b64, "mime": mime})
                else:
                    text = extract_text_from_file(file_data, file_rec["original_filename"], file_rec["content_type"])
                    if text:
                        file_text_parts.append(f"[File: {file_rec['original_filename']}]\n{text}")
            except Exception as e:
                logger.error(f"Error processing file {fid}: {e}")

    # Save user message with attachments
    user_msg = {
        "id": str(uuid.uuid4()),
        "chat_id": chat_id,
        "role": "user",
        "content": data.content,
        "attachments": attachments if attachments else None,
        "timestamp": now_ts,
    }
    await db.messages.insert_one({**user_msg, "_id": user_msg["id"]})

    # Auto-title on first message
    existing = await db.messages.count_documents({"chat_id": chat_id})
    if existing <= 1:
        title = data.content[:50] + ("..." if len(data.content) > 50 else "")
        if not data.content.strip() and attachments:
            title = f"[{attachments[0]['filename']}]"
        await db.chats.update_one({"id": chat_id}, {"$set": {"title": title}})

    # Get LLM response
    model_id = chat.get("model", "gpt-4o")
    provider = MODEL_PROVIDER_MAP.get(model_id, "openai")

    if provider == "openai_compatible" or (model_id.startswith("openai/") and provider not in EMERGENT_PROVIDERS):
        provider = "openai_compatible"

    conn = await db.connections.find_one({"_id": "global"}, {"_id": 0})
    default_key = os.environ.get("EMERGENT_LLM_KEY", "")

    if conn:
        prov_config = conn.get("providers", {}).get(provider, {})
        if prov_config.get("useEmergentKey", True) and provider in EMERGENT_PROVIDERS:
            api_key = default_key
        else:
            api_key = prov_config.get("apiKey", "")
        model_params = conn.get("modelParams", {})
    else:
        api_key = default_key
        model_params = {}

    temperature = data.temperature if data.temperature is not None else model_params.get("temperature", 0.7)
    max_tokens = data.max_tokens if data.max_tokens is not None else model_params.get("maxTokens", 4096)
    top_p = data.top_p if data.top_p is not None else model_params.get("topP", 1.0)

    # Build the user prompt with file context
    user_prompt = data.content or ""
    if file_text_parts:
        user_prompt = user_prompt + "\n\n" + "\n\n".join(file_text_parts) if user_prompt else "\n\n".join(file_text_parts)

    # Web search - automatically fetch relevant results
    search_context = ""
    if user_prompt.strip():
        search_context = web_search(user_prompt[:200], max_results=5)

    try:
        app_settings = await db.app_settings.find_one({"_id": "global"}, {"_id": 0})
        base_system_prompt = (app_settings or {}).get("systemPrompt", "You are a helpful AI assistant. You provide clear, accurate, and well-formatted responses using markdown when appropriate.")

        if search_context:
            system_prompt = base_system_prompt + "\n\nYou have access to recent web search results. Use them to provide up-to-date, accurate information. Cite sources when relevant using [Source](url) format.\n\n--- Web Search Results ---\n" + search_context + "\n--- End of Search Results ---"
        else:
            system_prompt = base_system_prompt

        # Build chat history
        history_msgs = await db.messages.find(
            {"chat_id": chat_id}, {"_id": 0}
        ).sort("timestamp", 1).to_list(50)

        if provider in EMERGENT_PROVIDERS:
            llm = LlmChat(
                api_key=api_key,
                session_id=chat_id,
                system_message=system_prompt
            ).with_model(provider, model_id)

            if image_b64_list:
                file_contents = [ImageContent(image_base64=img["b64"]) for img in image_b64_list]
                user_message = UserMessage(text=user_prompt or "What do you see in this image?", file_contents=file_contents)
            else:
                user_message = UserMessage(text=user_prompt)
            response_text = await llm.send_message(user_message)
        else:
            # Use litellm
            messages_for_llm = [{"role": "system", "content": system_prompt}]
            for hm in history_msgs:
                messages_for_llm.append({"role": hm["role"], "content": hm["content"]})

            # Build multimodal content for the current message if images present
            if image_b64_list:
                content_parts = []
                if user_prompt:
                    content_parts.append({"type": "text", "text": user_prompt})
                for img in image_b64_list:
                    content_parts.append({"type": "image_url", "image_url": {"url": f"data:{img['mime']};base64,{img['b64']}"}})
                messages_for_llm.append({"role": "user", "content": content_parts})
            else:
                messages_for_llm.append({"role": "user", "content": user_prompt})

            litellm_model = model_id
            extra_kwargs = {}

            if provider == "deepseek":
                os.environ["DEEPSEEK_API_KEY"] = api_key
            elif provider == "qwen":
                base_url = (conn or {}).get("providers", {}).get("qwen", {}).get("baseUrl", PROVIDER_BASE_URLS.get("qwen", ""))
                extra_kwargs["api_base"] = base_url
                extra_kwargs["api_key"] = api_key
            elif provider == "grok":
                os.environ["XAI_API_KEY"] = api_key
            elif provider == "perplexity":
                os.environ["PERPLEXITYAI_API_KEY"] = api_key
            elif provider == "bedrock":
                bedrock_config = (conn or {}).get("providers", {}).get("bedrock", {})
                os.environ["AWS_ACCESS_KEY_ID"] = bedrock_config.get("awsAccessKey", "")
                os.environ["AWS_SECRET_ACCESS_KEY"] = bedrock_config.get("awsSecretKey", "")
                os.environ["AWS_REGION_NAME"] = bedrock_config.get("awsRegion", "us-east-1")
            elif provider == "openai_compatible":
                compat_config = (conn or {}).get("providers", {}).get("openai_compatible", {})
                base_url = compat_config.get("baseUrl", "")
                extra_kwargs["api_base"] = base_url
                extra_kwargs["api_key"] = api_key

            response = await litellm.acompletion(
                model=litellm_model,
                messages=messages_for_llm,
                temperature=temperature,
                max_tokens=max_tokens,
                top_p=top_p,
                **extra_kwargs,
            )
            response_text = response.choices[0].message.content

    except Exception as e:
        logger.error(f"LLM error for provider={provider} model={model_id}: {e}")
        response_text = f"Sorry, I encountered an error while generating a response. Please try again.\n\nError: {str(e)}"

    ai_msg = {
        "id": str(uuid.uuid4()),
        "chat_id": chat_id,
        "role": "assistant",
        "content": response_text,
        "web_searched": bool(search_context),
        "timestamp": datetime.now(timezone.utc).isoformat(),
    }
    await db.messages.insert_one({**ai_msg, "_id": ai_msg["id"]})

    return {
        "user_message": {k: v for k, v in user_msg.items() if k != "_id"},
        "assistant_message": {k: v for k, v in ai_msg.items() if k != "_id"},
    }

# ===== Archive, Export, Import Endpoints =====

@api_router.put("/chats/{chat_id}/archive")
async def archive_chat(chat_id: str, user=Depends(get_current_user)):
    result = await db.chats.update_one({"id": chat_id, "user_id": user["id"]}, {"$set": {"archived": True}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Chat not found")
    return {"status": "ok"}

@api_router.put("/chats/{chat_id}/unarchive")
async def unarchive_chat(chat_id: str, user=Depends(get_current_user)):
    result = await db.chats.update_one({"id": chat_id, "user_id": user["id"]}, {"$set": {"archived": False}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Chat not found")
    return {"status": "ok"}

@api_router.get("/chats/{chat_id}/export")
async def export_chat(chat_id: str, user=Depends(get_current_user)):
    chat = await db.chats.find_one({"id": chat_id, "user_id": user["id"]}, {"_id": 0})
    if not chat:
        raise HTTPException(status_code=404, detail="Chat not found")
    messages = await db.messages.find({"chat_id": chat_id}, {"_id": 0}).sort("timestamp", 1).to_list(1000)
    return {
        "version": "1.0",
        "source": "open-webui-clone",
        "chat": chat,
        "messages": messages,
        "exported_at": datetime.now(timezone.utc).isoformat(),
    }

class ChatImport(BaseModel):
    version: str = "1.0"
    chat: dict
    messages: list

@api_router.post("/chats/import")
async def import_chat(data: ChatImport, user=Depends(get_current_user)):
    new_chat_id = str(uuid.uuid4())
    chat_data = {
        "id": new_chat_id,
        "title": data.chat.get("title", "Imported Chat"),
        "model": data.chat.get("model", "gpt-4o"),
        "user_id": user["id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "archived": False,
    }
    await db.chats.insert_one({**chat_data, "_id": new_chat_id})

    for msg in data.messages:
        new_msg = {
            "id": str(uuid.uuid4()),
            "chat_id": new_chat_id,
            "role": msg.get("role", "user"),
            "content": msg.get("content", ""),
            "timestamp": msg.get("timestamp", datetime.now(timezone.utc).isoformat()),
        }
        await db.messages.insert_one({**new_msg, "_id": new_msg["id"]})

    return {"status": "ok", "chat_id": new_chat_id, "chat": chat_data}

@api_router.delete("/chats/archived/all")
async def delete_all_archived(user=Depends(get_current_user)):
    archived = await db.chats.find({"user_id": user["id"], "archived": True}, {"id": 1, "_id": 0}).to_list(1000)
    for chat in archived:
        await db.messages.delete_many({"chat_id": chat["id"]})
    await db.chats.delete_many({"user_id": user["id"], "archived": True})
    return {"status": "ok", "deleted_count": len(archived)}

# ===== Settings Endpoints =====

@api_router.get("/settings")
async def get_settings():
    doc = await db.app_settings.find_one({"_id": "global"}, {"_id": 0})
    return doc or {}

@api_router.put("/settings")
async def update_settings(data: dict):
    await db.app_settings.update_one(
        {"_id": "global"},
        {"$set": data},
        upsert=True
    )
    return {"status": "ok"}

# Include the router in the main app after all routes are defined
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("startup")
async def startup_event():
    try:
        init_storage()
        logger.info("Object storage initialized successfully")
    except Exception as e:
        logger.error(f"Object storage init failed: {e}")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()